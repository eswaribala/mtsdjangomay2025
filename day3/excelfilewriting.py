import os
from datetime import datetime
import random
from faker import Faker
from gender import Gender
import openpyxl
from openpyxl.styles import Font
fake= Faker("en_IN")  # Initialize Faker with Hindi locale
from models.employee import Employee
employees= []

for _ in  range(1,10):
    
    employee=Employee(
        
        fake.name(),
        fake.name(),
        fake.random_int(min=30000, max=120000),          
        fake.email(),
        fake.phone_number(),
        fake.job(),
        fake.date_of_birth(minimum_age=20, maximum_age=60).strftime('%Y-%m-%d'),
        fake.boolean(),
        random.choice(list(Gender))
  
       
    )
    employees.append(employee)
sorted_employees = sorted(employees, key=lambda x: x.get_first_name(), reverse=False)

dirPath="E:\\MTSTraningmay2025\\day3\\reports"
if not os.path.exists(dirPath):
    os.makedirs(dirPath)

wb= openpyxl.Workbook()
#ws1 = wb.create_sheet("May-2025")
ws = wb.active  
ws.title = "Employee Report"
# Write the header  
ws.append(["first_name", "last_name", "email", "phone_number", "salary", "job_title","status","hire_date", "gender"])
# Write the employee data   
for employee in sorted_employees:  
    ws.append([
        employee.get_first_name(),
        employee.get_last_name(),       
        employee.get_email(),
        employee.get_phone_number(),
         employee.get_salary(),
        employee.get_job_title(),        
        employee.get_status(),
        employee.get_hire_date(), 
        employee.get_gender().name
    ])
#for row in ws.iter_rows():
    #for cell in row:
      #  cell.font =  Font(name='Arial', size=24, bold=True if cell.row == 1 else False)
filePath=os.path.join(dirPath, f"employee_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
wb.save(filePath)
wb.close()